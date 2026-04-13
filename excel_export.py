"""
excel_export.py – Génération de l'export Excel au format CPS.
Importé par extensions.py.

Fonctions publiques:
    build_provenance_string(works)  → str
    build_excel(rows)               → openpyxl.Workbook
"""

from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ─── Constante colonnes ───────────────────────────────────────────────────────

HEADERS = [
    "ANNEE",
    "Désignation des préstations",
    "Quantité/Unité",
    "Quantité/ml",
    "AGR",
    "Secteurs",
    "Réseaux",
    "Provenance fourniture (marché + séries)",
    "PRIX UNITAIRE DE FOURNITURE",
    "PRIX UNITAIRE DE POSE",
    "PRIX TOTAL DE FOURNITURE",
    "PRIX TOTAL DE POSE",
]

# Largeurs en caractères
COL_WIDTHS = [8, 24, 14, 14, 10, 14, 24, 46, 22, 22, 22, 22]

# Indices des colonnes numériques (1-based)
NUM_COLS = {3, 4, 9, 10, 11, 12}

# ─── Couleurs ────────────────────────────────────────────────────────────────

C_HDR_BG   = "1D4ED8"   # bleu header
C_HDR_FG   = "FFFFFF"
C_ROW_EVEN = "EFF6FF"   # bleu très clair
C_ROW_ODD  = "FFFFFF"
C_TOT_BG   = "DBEAFE"   # bleu total
C_TOT_FG   = "1E3A8A"
C_TITLE_BG = "EFF6FF"
C_TITLE_FG = "1D4ED8"

# ─── Helpers styles ───────────────────────────────────────────────────────────

def _thin_border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _medium_border(color="1D4ED8"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ─── Chaîne de provenance ─────────────────────────────────────────────────────

def build_provenance_string(works):
    """
    Construit la chaîne de provenance CPS pour un groupe de travaux.

    Règles :
      - Regrouper par marché (reference)
      - Séries : dédupliquées, triées numériquement (sinon alphabétique), jointes par '-'
      - Marchés : triés alphabétiquement, joints par '; '

    Exemple : "M13/23: 40-64-122; M11/22: 95-101-102"

    Args:
        works: list[dict] – chaque dict contient 'marche' et 'serie'.

    Returns:
        str – chaîne de provenance formatée.
    """
    marche_series = defaultdict(set)

    for w in works:
        ref   = (w.get("marche") or "").strip()
        serie = str(w.get("serie") or "").strip()
        if not ref:
            continue
        if serie:
            marche_series[ref].add(serie)
        else:
            # garder le marché même sans série
            marche_series.setdefault(ref, set())

    if not marche_series:
        return ""

    parts = []
    for ref in sorted(marche_series.keys()):
        series = marche_series[ref]
        if series:
            def _key(s):
                try:    return (0, int(s))
                except: return (1, s)
            sorted_s = sorted(series, key=_key)
            parts.append(f"{ref}: {'-'.join(sorted_s)}")
        else:
            parts.append(ref)

    return "; ".join(parts)

# ─── Construction du Workbook ─────────────────────────────────────────────────

def build_excel(rows):
    """
    Génère un fichier Excel au format CPS à partir d'une liste de travaux approuvés.

    Args:
        rows: list[dict] – travaux avec clés :
              annee, diametre, agr, secteur, reseau,
              quantite_unite, quantite_ml, marche, serie,
              prix_four, prix_pose_u

    Returns:
        openpyxl.Workbook prêt à sauvegarder.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Export CPS"
    ws.sheet_view.showGridLines = False

    # ── Ligne de titre ────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    tc = ws["A1"]
    tc.value     = "SUIVI DES TRAVAUX – FORMAT CPS (travaux approuvés uniquement)"
    tc.font      = Font(bold=True, size=13, color=C_TITLE_FG)
    tc.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── En-têtes ──────────────────────────────────────────────────────────────
    for col, header in enumerate(HEADERS, 1):
        c           = ws.cell(row=2, column=col, value=header)
        c.font      = Font(bold=True, color=C_HDR_FG, size=9)
        c.fill      = PatternFill("solid", fgColor=C_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _thin_border("FFFFFF")
    ws.row_dimensions[2].height = 38
    ws.freeze_panes = "A3"

    # Largeurs
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Regroupement ──────────────────────────────────────────────────────────
    # Clé : (annee, agr, secteur, reseau, diametre)
    groups = defaultdict(list)
    for r in rows:
        key = (r["annee"], r["agr"], r["secteur"], r["reseau"], r["diametre"])
        groups[key].append(r)

    # ── Données ───────────────────────────────────────────────────────────────
    data_row = 3
    grand    = dict(unite=0.0, ml=0.0, four=0.0, pose=0.0)

    for idx, (key, works) in enumerate(sorted(groups.items())):
        annee, agr, secteur, reseau, diametre = key

        total_unite = sum(w["quantite_unite"] for w in works)
        total_ml    = sum(w["quantite_ml"]    for w in works)
        total_four  = sum(w["quantite_ml"] * w["prix_four"]   for w in works)
        total_pose  = sum(w["quantite_ml"] * w["prix_pose_u"] for w in works)

        # Prix unitaire unique seulement si constant dans le groupe
        four_set = {w["prix_four"]   for w in works}
        pose_set = {w["prix_pose_u"] for w in works}
        prix_four_u = list(four_set)[0] if len(four_set) == 1 else None
        prix_pose_u = list(pose_set)[0] if len(pose_set) == 1 else None

        prov_str = build_provenance_string(works)

        ws.append([
            annee,
            diametre,
            round(total_unite, 4),
            round(total_ml,    4),
            agr,
            secteur,
            reseau,
            prov_str,
            prix_four_u,
            prix_pose_u,
            round(total_four, 2),
            round(total_pose, 2),
        ])

        # Style de la ligne
        fill_color = C_ROW_EVEN if idx % 2 == 0 else C_ROW_ODD
        fill       = PatternFill("solid", fgColor=fill_color)
        border     = _thin_border()

        for col in range(1, len(HEADERS) + 1):
            c        = ws.cell(row=data_row, column=col)
            c.fill   = fill
            c.font   = Font(size=9)
            c.border = border
            if col in NUM_COLS:
                c.alignment    = Alignment(horizontal="right")
                c.number_format = "#,##0.00"
            elif col == 1:
                c.alignment = Alignment(horizontal="center")
            else:
                c.alignment = Alignment(horizontal="left", wrap_text=(col == 8))

        ws.row_dimensions[data_row].height = 16

        grand["unite"] += total_unite
        grand["ml"]    += total_ml
        grand["four"]  += total_four
        grand["pose"]  += total_pose
        data_row += 1

    # ── Ligne totaux ──────────────────────────────────────────────────────────
    ws.append([
        "TOTAL", "",
        round(grand["unite"], 4),
        round(grand["ml"],    4),
        "", "", "", "", "", "",
        round(grand["four"], 2),
        round(grand["pose"], 2),
    ])
    tot_border = _medium_border()
    for col in range(1, len(HEADERS) + 1):
        c           = ws.cell(row=data_row, column=col)
        c.fill      = PatternFill("solid", fgColor=C_TOT_BG)
        c.font      = Font(bold=True, size=10, color=C_TOT_FG)
        c.border    = tot_border
        c.alignment = Alignment(
            horizontal="right" if col in NUM_COLS else
            "center" if col == 1 else "left"
        )
        if col in NUM_COLS:
            c.number_format = "#,##0.00"
    ws.row_dimensions[data_row].height = 20

    return wb
